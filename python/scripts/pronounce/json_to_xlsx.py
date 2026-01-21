import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


# =============================
# 🔍 1. Kiểm tra dữ liệu JSON
# =============================
def validate_json_data(data):
    if not isinstance(data, dict):
        raise ValueError("JSON không đúng định dạng: phải là một dictionary.")
    if "data" not in data:
        raise ValueError("JSON thiếu khóa 'data'.")
    errors = []
    for q_index, item in enumerate(data.get("data", []), start=1):
        if not isinstance(item, list) or len(item) != 3:
            errors.append(f"Câu {q_index}: Cấu trúc không hợp lệ, cần [words, phonetics, answer_index].")
            continue
        words, phonetics, answer_index = item
        if not isinstance(words, list) or not isinstance(phonetics, list):
            errors.append(f"Câu {q_index}: 'words' hoặc 'phonetics' không phải là danh sách.")
            continue
        if len(words) != len(phonetics):
            errors.append(f"Câu {q_index}: Số từ ({len(words)}) không khớp với số phiên âm ({len(phonetics)}).")
            continue
        if not isinstance(answer_index, int) or answer_index < 0 or answer_index >= len(words):
            errors.append(f"Câu {q_index}: 'answer_index' không hợp lệ: {answer_index}.")
            continue
    if errors:
        print("Các lỗi trong dữ liệu JSON:")
        for error in errors:
            print(error)
        return False
    return True


# =============================
# 🧼 2. Xử lý tên sheet Excel
# =============================
def sanitize_sheet_name(name: str) -> str:
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in invalid_chars:
        name = name.replace(ch, "–")
    return name[:31]


# =============================
# 📊 3. Chuyển JSON → Excel
# =============================
def convert_json_to_xlsx():
    base_dir = Path(__file__).resolve().parents[2]
    json_path = base_dir / "data/pronounce/json/pronounce-set-2.json"
    output_dir = base_dir / "data/pronounce/xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Đọc file JSON
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"Lỗi khi đọc file JSON: {e}")
        return

    # Kiểm tra dữ liệu
    if not validate_json_data(data):
        print("Dừng xử lý do lỗi dữ liệu JSON.")
        return

    title = sanitize_sheet_name(data.get("title", "Pronunciation Set"))
    sounds = data.get("sounds", [])
    all_rows = []
    seen_pairs = set()

    # === Chuyển từng câu hỏi thành bảng dữ liệu ===
    for q_index, item in enumerate(data.get("data", []), start=1):
        words, phonetics, answer_index = item
        for i, (word, phonetic) in enumerate(zip(words, phonetics)):
            pair_key = (q_index, word, phonetic)
            if pair_key in seen_pairs:
                print(f"Câu {q_index}: Bỏ qua cặp trùng lặp: {word}, {phonetic}")
                continue
            seen_pairs.add(pair_key)
            all_rows.append({
                "Question No.": q_index,
                "Word Option": word,
                "Phonetic": phonetic,
                "Correct?": i == answer_index,
            })

    if not all_rows:
        print("Không có dữ liệu hợp lệ để ghi ra Excel.")
        return

    # === Tạo DataFrame chính (không lặp Sound 1/2) ===
    df = pd.DataFrame(all_rows)[["Question No.", "Word Option", "Phonetic", "Correct?"]]

    output_path = output_dir / f"{json_path.stem}.xlsx"
    temp_path = output_dir / f"_{json_path.stem}_temp.xlsx"

    # Ghi dữ liệu tạm (bắt đầu từ hàng 4 để chừa phần Sound)
    df.to_excel(temp_path, index=False, startrow=3)

    # === Mở Excel và ghi thêm thông tin Sound ===
    wb = load_workbook(temp_path)
    ws = wb.active
    ws.title = title

    # Hai dòng đầu: Sound 1 và Sound 2
    ws["A1"] = "Sound 1"
    ws["B1"] = sounds[0] if len(sounds) > 0 else ""
    ws["A2"] = "Sound 2"
    ws["B2"] = sounds[1] if len(sounds) > 1 else ""

    # Ghi header bảng dữ liệu ở hàng 3
    ws["A3"] = "Question No."
    ws["B3"] = "Word Option"
    ws["C3"] = "Phonetic"
    ws["D3"] = "Correct?"

    # === Thêm dòng trống sau mỗi 4 dòng câu hỏi ===
    insert_offset = 0
    start_row = 10  # dữ liệu bắt đầu ở đây
    for row in range(start_row + 4, ws.max_row + insert_offset + 2, 5):
        ws.insert_rows(row + insert_offset)
        insert_offset += 1

    wb.save(output_path)
    temp_path.unlink(missing_ok=True)

    # === In thống kê ===
    print(f"✅ File Excel đã được tạo: {output_path}")
    print(f"📘 Sheet: {title}")
    print(f"🔉 Sound 1: {sounds[0] if sounds else '(không có)'}")
    print(f"🔉 Sound 2: {sounds[1] if len(sounds) > 1 else '(không có)'}")
    print(f"📊 Tổng số câu hỏi: {len(data.get('data', []))}")
    print(f"📄 Tổng số dòng (kể cả dòng trống): {ws.max_row}")


# =============================
# 🚀 4. Chạy script
# =============================
if __name__ == "__main__":
    convert_json_to_xlsx()
